# -*- coding: utf-8 -*-
"""Extract fabric data from Excel files into a unified JSON schema.

Reads the source files under 面料推荐档案/ (read-only) and writes three JSON
artefacts under data/:

    fabrics.json   - unified fabric catalogue
    suppliers.json - supplier directory derived from the catalogue
    styles.json    - garment-style notes (3S-AVVA 2026.3.9 xlsx/csv)

Run:
    python scripts/extract_fabrics.py            # write JSON
    python scripts/extract_fabrics.py --dry-run  # preview counts only
    python scripts/extract_fabrics.py --source <name>  # extract one source
"""
from __future__ import annotations

import argparse
import glob
import hashlib
import json
import re
import sys

# Make stdout unicode-safe on Windows terminals that default to GBK
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "面料推荐档案"
OUTPUT_DIR = ROOT / "data"

# ---------------------------------------------------------------------------
# Canonical fabric name dictionary (zh/en/abbrev -> canonical english key)
# ---------------------------------------------------------------------------
FIBER_MAP: dict[str, str] = {
    # polyester family
    "涤纶": "polyester", "聚酯": "polyester", "POLYESTER": "polyester", "PES": "polyester",
    "再生涤": "recycled_polyester",
    # spandex / elastane
    "氨纶": "spandex", "弹性纤维": "spandex", "弹性烷": "spandex",
    "SP": "spandex", "SPANDEX": "spandex", "ELASTANE": "spandex",
    # cotton
    "棉": "cotton", "全棉": "cotton", "COTTON": "cotton",
    # nylon
    "锦纶": "nylon", "尼龙": "nylon", "NYLON": "nylon", "NL": "nylon",
    # rayon / modal / viscose
    "粘胶": "rayon", "粘胶纤维": "rayon", "人棉": "rayon", "VISCOSE": "rayon",
    "莫代尔": "modal", "莫达尔": "modal", "MODAL": "modal",
    # linen
    "亚麻": "linen", "LINEN": "linen",
    # acrylic
    "腈纶": "acrylic", "ACRYLIC": "acrylic", "Acrylic": "acrylic",
}

CATEGORIES = {
    "knit": "针织",
    "woven": "化纤长丝/梭织",
    "pu_suede": "PU/麂皮",
    "home_textile": "家纺阻燃",
}

# ---------------------------------------------------------------------------
# Field helpers
# ---------------------------------------------------------------------------
PCT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*%?\s*([A-Za-z一-鿿]{1,12})")
NON_FIBER_WORDS = {"gsm", "g/m²", "g/m2", "white", "black", "coating", "linen", "like", "yarn", "yarns", "bone"}


def parse_composition(text: Any) -> dict[str, float]:
    """Parse composition strings into {fiber: percent}.

    Handles:
      - "94%涤纶 + 6%氨纶"
      - "49%粘胶纤维48%聚酯3%弹性纤维" (no separators)
      - "92% NYLON、8% SP"
      - "57% PES / 3% Acrylic, 320gsm, white"  # strips 'gsm', 'white'

    Tokens that are clearly not fibres are dropped (see NON_FIBER_WORDS).
    """
    if not text:
        return {}
    cleaned = (
        str(text)
        .replace("，", " ")
        .replace("、", " ")
        .replace("+", " ")
        .replace("\n", " ")
        .replace(",", " ")
    )
    result: dict[str, float] = {}
    for pct, name in PCT_RE.findall(cleaned):
        try:
            value = float(pct)
        except ValueError:
            continue
        clean_name = name.strip().lower()
        if clean_name in NON_FIBER_WORDS:
            continue
        key = (
            FIBER_MAP.get(name)
            or FIBER_MAP.get(name.strip().upper())
            or clean_name
        )
        result[key] = round(result.get(key, 0.0) + value, 2)
    return result


def parse_width(text: Any) -> float | None:
    """Normalise width to centimetres. Accepts '150cm', '57/58英寸', '280cm'."""
    if not text:
        return None
    t = str(text).strip()
    m = re.search(r"(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)\s*(英寸|in|inch|\")?", t, re.I)
    if m:
        a, b = float(m.group(1)), float(m.group(2))
        unit = (m.group(3) or "cm").lower()
        avg = (a + b) / 2
        if unit.startswith(("in", "英", '"')):
            return round(avg * 2.54, 1)
        return round(avg, 1)
    m = re.search(r"(\d+(?:\.\d+)?)\s*(cm|厘米|英寸|in|inch|\")", t, re.I)
    if m:
        val, unit = float(m.group(1)), m.group(2).lower()
        if unit.startswith(("in", "英", '"')):
            return round(val * 2.54, 1)
        return val
    return None


def parse_weight(text: Any) -> tuple[float | None, dict[str, int] | None]:
    """Return (gsm, range_dict). Prefers explicit g/m² token over width digits.

    Recognised weight units: 'gsm', 'g/m²', 'g/m2', 'g/㎡' (CJK U+33A1).
    The pattern is '<digits><space?>g<unit>'; for '165cm×380g/㎡' the function
    finds '380g/㎡' and returns 380.
    """
    if text is None or text == "":
        return None, None
    t = str(text).strip()

    # Range like '180-200 g/m²'
    m = re.match(r"^(\d+)\s*[-–~]\s*(\d+)\s*g", t, re.I)
    if m:
        return (int(m.group(1)) + int(m.group(2))) / 2, {"min": int(m.group(1)), "max": int(m.group(2))}

    # '380g/㎡', '380 g/m²', '380 g/m2', '380 gsm'
    m = re.search(r"(\d+)\s*g\s*(?:/|／|sm|m²|㎡|m2)", t, re.I)
    if m:
        return int(m.group(1)), None

    # Bare '380gsm' (no slash)
    m = re.search(r"(\d+)\s*gsm", t, re.I)
    if m:
        return int(m.group(1)), None

    # Range without unit, e.g. '180-200'
    m = re.match(r"^(\d+)\s*[-–~]\s*(\d+)$", t)
    if m:
        return (int(m.group(1)) + int(m.group(2))) / 2, {"min": int(m.group(1)), "max": int(m.group(2))}

    nums = re.findall(r"\d+", t)
    if nums:
        return int(nums[0]), None
    return None, None


def clean_supplier_name(raw: Any) -> str:
    """Repair supplier names split by stray commas (e.g. '绍兴金典 , , , ,进出口有限公司')."""
    if raw is None:
        return ""
    s = str(raw)
    s = re.sub(r"(\s*[,，]\s*){2,}", "", s)
    s = re.sub(r"\s+", "", s)
    return s.strip()


def make_id(*parts: Any) -> str:
    seed = "|".join("" if p is None else str(p) for p in parts).encode("utf-8")
    return hashlib.md5(seed).hexdigest()[:12]


# ---------------------------------------------------------------------------
# Source extractors
# ---------------------------------------------------------------------------
def extract_huarui(src: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    items: list[dict[str, Any]] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        seq, name, code, comp, spec, features, uses = (list(row) + [None] * 7)[:7]
        if not name:
            continue
        weight, weight_range = parse_weight(spec)
        items.append({
            "id": make_id("huarui", code, name),
            "supplier": "常熟市华瑞针纺织",
            "category": "knit",
            "name": str(name).strip(),
            "code": str(code).strip() if code else None,
            "composition_raw": comp,
            "composition": parse_composition(comp),
            "spec_raw": spec,
            "weight_gsm": weight,
            "weight_range": weight_range,
            "features": [str(features).strip()] if features else [],
            "applications": [a.strip() for a in str(uses).replace("、", ",").split(",") if a.strip()] if uses else [],
            "source_file": str(src.relative_to(ROOT)),
            "source_row": row_no,
        })
    return items


def extract_zhongtao(src: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    items: list[dict[str, Any]] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, comp, weave, width, weight, structure, finish, price, uses = (list(row) + [None] * 10)[:10]
        if not code:
            continue
        items.append({
            "id": make_id("zhongtao", code, name),
            "supplier": "中涛三时",
            "category": "woven",
            "name": str(name).strip() if name else None,
            "code": str(code).strip(),
            "composition_raw": comp,
            "composition": parse_composition(comp),
            "weave": weave,
            "structure": structure,
            "finish": finish,
            "spec_raw": width,
            "width_cm": parse_width(width),
            "weight_gsm": parse_weight(weight)[0],
            "price_rmb_per_m": price,
            "applications": [a.strip() for a in re.split(r"[\s,,]+", str(uses)) if a.strip()] if uses else [],
            "tags": [],
            "source_file": str(src.relative_to(ROOT)),
            "source_row": row_no,
        })
    return items


def extract_wantai(src: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    items: list[dict[str, Any]] = []
    # Row 1 has empty col A then headers from col B; data starts row 2.
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        seq, name, texture, color = (list(row) + [None] * 5)[1:5]
        if not name:
            continue
        items.append({
            "id": make_id("wantai", seq, name),
            "supplier": "万泰",
            "category": "pu_suede",
            "name": str(name).strip(),
            "code": str(seq).strip() if seq else None,
            "texture": texture,
            "color": color,
            "tags": [t.strip() for t in re.split(r"[、,，]+", str(texture)) if t.strip()] if texture else [],
            "source_file": str(src.relative_to(ROOT)),
            "source_row": row_no,
        })
    return items


def extract_home_fr(src: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    items: list[dict[str, Any]] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        hb, width, name, fr_std, fabric, edge, moq, fob = (list(row) + [None] * 8)[:8]
        if not name:
            continue
        items.append({
            "id": make_id("home_fr", hb, name),
            "supplier": "杭州辦事處",
            "category": "home_textile",
            "name": str(name).strip(),
            "code": str(hb).strip() if hb else None,
            "composition_raw": fabric,
            "composition": parse_composition(fabric),
            "flame_retardant": True,
            "fr_standard": fr_std,
            "edge": edge,
            "spec_raw": width,
            "width_cm": parse_width(width),
            "moq": moq,
            "fob_usd_per_m": fob,
            "source_file": str(src.relative_to(ROOT)),
            "source_row": row_no,
        })
    return items


def extract_3savva(src: Path) -> list[dict[str, Any]]:
    """The 3S-AVVA sheet puts fabric info on the first row, then 1 row per supplier."""
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    items: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row_no, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        seq, name, comp, spec, weight, supplier, price, moq, phone, email, note = (list(row) + [None] * 11)[:11]
        if name:
            weight_val, weight_range = parse_weight(weight)
            current = {
                "id": make_id("3savva", seq, name),
                "supplier_brand": "3S-AVVA",
                "category": "knit",
                "name": str(name).strip(),
                "code": str(seq).strip() if seq else None,
                "composition_raw": comp,
                "composition": parse_composition(comp),
                "spec_raw": spec,
                "weight_gsm": weight_val,
                "weight_range": weight_range,
                "supplier_quotes": [],
                "source_file": str(src.relative_to(ROOT)),
                "source_row_first": row_no,
            }
            items.append(current)
        if current and supplier:
            current["supplier_quotes"].append({
                "supplier": clean_supplier_name(supplier),
                "price_rmb_per_m": price,
                "moq": moq,
                "phone": phone,
                "email": email,
            })
    return items


def extract_styles(src_xlsx: Path, src_csv: Path) -> list[dict[str, Any]]:
    """3S-AVVA 2026.3.9 xlsx/csv contains garment-style notes, not pure fabric data.

    Layout (xlsx): col A is empty, data lives in cols B-E and col L.
        B = 款式描述 (style description)
        C = 面料描述 (fabric description / composition)
        L = extra notes (rare)
    The companion CSV uses the same logical columns without the empty col A.
    """
    items: list[dict[str, Any]] = []
    wb = openpyxl.load_workbook(src_xlsx, data_only=True)
    ws = wb.active
    seq = 0
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Style description lives in column B (index 1); col A is always empty.
        style = row[1] if len(row) > 1 else None
        if not style:
            continue
        seq += 1
        fabric_desc = row[2] if len(row) > 2 else None
        extra = row[11] if len(row) > 11 else None
        items.append({
            "id": make_id("3savva_style", seq, style),
            "supplier_brand": "3S-AVVA",
            "category": "knit_style",
            "style_description": str(style).strip(),
            "fabric_description": str(fabric_desc).strip() if fabric_desc else None,
            "fabric_composition": parse_composition(fabric_desc) if fabric_desc else {},
            "extra_notes": str(extra).strip() if extra else None,
            "source_file": str(src_xlsx.relative_to(ROOT)),
            "source_row": row_no,
        })
    return items


# ---------------------------------------------------------------------------
# Sources registry
# ---------------------------------------------------------------------------
def _resolve_source(rel: str) -> Path:
    """Glob-based resolver: tolerates non-breaking spaces / encoding drift in filenames."""
    matches = glob.glob(str(SOURCE_DIR / rel))
    if not matches:
        raise FileNotFoundError(f"Source file not found: {SOURCE_DIR / rel}")
    return Path(matches[0])


@dataclass
class Source:
    key: str
    rel_path: str
    extractor: Callable[[Path], list[dict[str, Any]]]

    def path(self) -> Path:
        return _resolve_source(self.rel_path)


SOURCES: list[Source] = [
    Source("huarui",   "面料推荐/华瑞针织面料汇总表.xlsx",            extract_huarui),
    Source("zhongtao", "面料推荐/中涛 三时.xlsx",                   extract_zhongtao),
    Source("wantai",   "面料推荐/万泰推荐面料.xlsx",                 extract_wantai),
    Source("home_fr",  "家纺/260407*收到面料.xlsx",                  extract_home_fr),
    Source("3savva",   "家纺/3S-AVVA_针织面料报价单_多供应商版_新增网址.xlsx", extract_3savva),
]


def derive_suppliers(fabrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Collapse supplier names + quote counts into a directory."""
    bucket: dict[str, dict[str, Any]] = {}
    for f in fabrics:
        for q in f.get("supplier_quotes") or []:
            name = q.get("supplier")
            if not name:
                continue
            entry = bucket.setdefault(
                name,
                {"name": name, "phone": q.get("phone"), "email": q.get("email"),
                 "quote_count": 0, "fabric_ids": []},
            )
            entry["quote_count"] += 1
            entry["fabric_ids"].append(f["id"])
        main = f.get("supplier") or f.get("supplier_brand")
        if main:
            entry = bucket.setdefault(
                main,
                {"name": main, "phone": None, "email": None,
                 "quote_count": 0, "fabric_ids": []},
            )
            entry["fabric_ids"].append(f["id"])
    for entry in bucket.values():
        entry["fabric_ids"] = sorted(set(entry["fabric_ids"]))
        entry["fabric_count"] = len(entry["fabric_ids"])
    return sorted(bucket.values(), key=lambda e: e["name"])


def main() -> int:
    ap = argparse.ArgumentParser(description="Build unified fabric catalogue from Excel sources.")
    ap.add_argument("--dry-run", action="store_true", help="Print counts without writing JSON.")
    ap.add_argument("--source", choices=[s.key for s in SOURCES] + ["styles"], help="Extract only one source.")
    args = ap.parse_args()

    selected = SOURCES if not args.source else [s for s in SOURCES if s.key == args.source]
    fabrics: list[dict[str, Any]] = []
    counts: dict[str, int] = {}
    for src in selected:
        items = src.extractor(src.path())
        fabrics.extend(items)
        counts[src.key] = len(items)
        print(f"  [{src.key}] extracted {len(items):>3} items from {src.path().name}")

    if args.dry_run:
        print("\\nDry run - nothing written.")
        print(json.dumps(counts, ensure_ascii=False, indent=2))
        return 0

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fabric_payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "schema_version": 1,
        "counts": counts,
        "total": len(fabrics),
        "fabrics": fabrics,
    }
    (OUTPUT_DIR / "fabrics.json").write_text(
        json.dumps(fabric_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  wrote {OUTPUT_DIR / 'fabrics.json'} ({len(fabrics)} items)")

    suppliers = derive_suppliers(fabrics)
    (OUTPUT_DIR / "suppliers.json").write_text(
        json.dumps({
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "total": len(suppliers),
            "suppliers": suppliers,
        }, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"  wrote {OUTPUT_DIR / 'suppliers.json'} ({len(suppliers)} suppliers)")

    if args.source in (None, "styles"):
        styles_path = SOURCE_DIR / "家纺" / "3S-AVVA 针织面料报价-2026.3.9.xlsx"
        if styles_path.exists():
            styles = extract_styles(styles_path, styles_path)
            (OUTPUT_DIR / "styles.json").write_text(
                json.dumps({
                    "generated_at": datetime.now().isoformat(timespec="seconds"),
                    "total": len(styles),
                    "styles": styles,
                }, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"  wrote {OUTPUT_DIR / 'styles.json'} ({len(styles)} style notes)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
