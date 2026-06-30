"""
数据加载：从 xlsx/csv 读出 86 条面料，转成钉钉 API 格式
"""
import logging
from pathlib import Path
import openpyxl

logger = logging.getLogger("dingtalk.data")


# xlsx 列名 → 钉钉 field alias
COLUMN_MAP = {
    "面料编号": "code",
    "面料名称": "name",
    "面料图片": "image",
    "品类": "category",
    "供应商": "supplier",
    "成分描述": "composition",
    "涤纶 %": "polyester_pct",
    "棉 %": "cotton_pct",
    "氨纶 %": "spandex_pct",
    "再生纤维 %": "recycled_pct",
    "其他成分": "other_composition",
    "幅宽 cm": "width_cm",
    "克重 g/㎡": "weight_gsm",
    "结构/织法": "structure",
    "后整理": "finish",
    "阻燃等级": "fr_standard",
    "RMB 价格 元/m": "price_rmb_per_m",
    "FOB 价格 USD/m": "price_fob_usd_per_m",
    "起订量 MOQ": "moq",
    "适用季节": "season",
    "推荐款式": "applications",
    "特性标签": "features",
    "卖点文案": "selling_points",
    "相似款链接": "similar_links",
    "备注": "remark",
    "__来源文件": "source_file",
    "__来源行号": "source_row",
    "__导入时间": "imported_at",
}

# 多选用分号分隔，需要拆成 list
MULTI_SELECT_ALIASES = {"season", "applications", "features", "finish"}

# 单选需要严格匹配字典值，否则钉钉会报错
SINGLE_SELECT_ALIASES = {"category", "supplier", "structure"}


def load_records_from_xlsx(xlsx_path: Path) -> list[dict]:
    """
    读取 xlsx，返回钉钉 API 格式的 records 列表
    [
        {"fields": {"code": "ML001", "name": "...", "category": "针织", ...}},
        ...
    ]
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"数据文件不存在：{xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 读取表头
    headers = [c.value for c in ws[1]]
    col_index = {h: i for i, h in enumerate(headers) if h}

    records = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        fields = {}
        for cn, alias in COLUMN_MAP.items():
            if cn not in col_index:
                continue
            v = row[col_index[cn]]
            if v is None or v == "":
                continue
            # 多选拆 list
            if alias in MULTI_SELECT_ALIASES and isinstance(v, str):
                v = [s.strip() for s in v.split(";") if s.strip()]
            # 单选转为字符串（如果非字典值会跳过，在校验时警告）
            if alias in SINGLE_SELECT_ALIASES:
                v = str(v).strip()
            # 数字转 int/float
            if alias in ("polyester_pct", "cotton_pct", "spandex_pct", "recycled_pct",
                         "width_cm", "weight_gsm", "price_rmb_per_m", "price_fob_usd_per_m",
                         "moq", "source_row") and v != "":
                try:
                    v = float(v) if "." in str(v) else int(v)
                except (ValueError, TypeError):
                    pass
            fields[alias] = v

        # 必填校验
        if not fields.get("code") or not fields.get("name"):
            skipped += 1
            logger.warning(f"  跳过一行（缺 code 或 name）: {row[:3]}")
            continue

        records.append({"fields": fields})

    logger.info(f"✅ 读取 {len(records)} 条记录（跳过 {skipped} 条）")
    return records


def load_records_from_csv(csv_path: Path) -> list[dict]:
    """从 CSV 读取，逻辑同上"""
    import csv
    if not csv_path.exists():
        raise FileNotFoundError(f"数据文件不存在：{csv_path}")

    records = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            fields = {}
            for cn, alias in COLUMN_MAP.items():
                v = row.get(cn, "").strip() if row.get(cn) else ""
                if not v:
                    continue
                if alias in MULTI_SELECT_ALIASES:
                    v = [s.strip() for s in v.split(";") if s.strip()]
                if alias in SINGLE_SELECT_ALIASES:
                    v = str(v).strip()
                if alias in ("polyester_pct", "cotton_pct", "spandex_pct", "recycled_pct",
                             "width_cm", "weight_gsm", "price_rmb_per_m", "price_fob_usd_per_m",
                             "moq", "source_row") and v:
                    try:
                        v = float(v) if "." in v else int(v)
                    except (ValueError, TypeError):
                        pass
                fields[alias] = v
            if not fields.get("code") or not fields.get("name"):
                continue
            records.append({"fields": fields})
    logger.info(f"✅ 读取 {len(records)} 条记录（CSV）")
    return records
