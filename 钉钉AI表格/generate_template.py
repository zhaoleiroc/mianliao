"""
生成：
1) 钉钉 AI 表格导入模板（含 1 条样例 + 字典页 + 字段说明页）
2) 从现有 86 条面料数据预转出的导入数据文件
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = r"C:\Users\zhaolei\Desktop\学习\面料推荐\钉钉AI表格"
SRC_JSON = r"C:\Users\zhaolei\Desktop\学习\面料推荐\data\fabrics.json"

# ---------- 1. 25 个字段定义 ----------
COLUMNS = [
    # (字段名, 钉钉类型, 必填, 字段说明, 示例值)
    ("面料编号",   "文本（唯一）", "✅", "内部唯一编号/货号",                  "ML-001"),
    ("面料名称",   "文本",        "✅", "中文品名，双面弹力摇粒绒等",          "双面弹力摇粒绒"),
    ("面料图片",   "附件/图片",   "📎", "拍实物图，AI 自动识别成分/克重",     ""),
    ("品类",       "单选",        "✅", "针织/机织/PU 绒/家纺",               "针织"),
    ("供应商",     "单选",        "✅", "5 家供应商之一",                     "常熟市华瑞针纺织"),
    ("成分描述",   "文本",        "✅", "如：94%涤纶 + 6%氨纶",                "94%涤纶 + 6%氨纶"),
    ("涤纶 %",     "数字",        "📎", "0-100",                              94),
    ("棉 %",       "数字",        "📎", "0-100",                              0),
    ("氨纶 %",     "数字",        "📎", "0-100",                              6),
    ("再生纤维 %", "数字",        "📎", "0-100",                              0),
    ("其他成分",   "文本",        "📎", "如：莫代尔/羊毛/锦纶",               ""),
    ("幅宽 cm",    "数字",        "📎", "常用 145-180",                       165),
    ("克重 g/㎡",  "数字",        "✅", "面料厚度关键指标",                    380),
    ("结构/织法",  "单选",        "📎", "见字典",                             "纬编针织"),
    ("后整理",     "多选",        "📎", "见字典，多选用分号分隔",             "磨毛"),
    ("阻燃等级",   "文本",        "📎", "如：NFPA 701 / EN 13501-1",          ""),
    ("RMB 价格 元/m", "数字",     "📎", "人民币每米报价",                     35),
    ("FOB 价格 USD/m", "数字",    "📎", "出口每米美元价",                     ""),
    ("起订量 MOQ", "数字",        "📎", "最小起订米数",                       500),
    ("适用季节",   "多选",        "✅", "春/夏/秋/冬/四季，多选用分号分隔",   "秋;冬"),
    ("推荐款式",   "多选",        "✅", "见字典，多选用分号分隔",             "外套;卫衣;家居服"),
    ("特性标签",   "多选",        "📎", "见字典，多选用分号分隔",             "保暖;弹力"),
    ("卖点文案",   "长文本",      "📎", "AI 自动生成 30-80 字推荐话术",        ""),
    ("相似款链接", "关联记录",    "📎", "AI 找同品类/克重接近的替代款",       ""),
    ("备注",       "长文本",      "📎", "自由备注",                           ""),
]

# 隐藏元数据
META_COLUMNS = [
    ("__来源文件", "文本", "溯源到原 Excel"),
    ("__来源行号", "数字", "溯源到原始行"),
    ("__导入时间", "自动日期", "第一次导入时间"),
]

# ---------- 字典 ----------
DICTS = {
    "品类": ["针织", "机织", "PU 绒", "家纺"],
    "供应商": ["常熟市华瑞针纺织", "中涛 / 三时", "万泰", "home_fr", "3S AVVA"],
    "结构/织法": ["纬编针织", "经编针织", "平纹", "斜纹", "缎纹", "提花", "灯芯绒", "摇粒绒复合", "PU 涂层"],
    "适用季节": ["春", "夏", "秋", "冬", "四季"],
    "推荐款式": ["外套", "T 恤", "卫衣", "衬衫", "夹克/夹棉", "家居服", "运动服", "童装", "内裤", "家纺床品", "家纺窗帘", "马甲", "工装裤", "Polo 衫"],
    "特性标签": ["保暖", "透气", "弹力", "亲肤", "挺括", "垂感", "抗皱", "抗起球", "防水", "阻燃", "再生环保", "复古"],
    "后整理": ["磨毛", "压花", "烫金", "烫银", "印花", "PU 涂层", "防水涂层", "抗静电", "抗菌", "阻燃处理"],
}

# ---------- 2. 推断季节 / 款式 / 特性 ----------
def infer_season(weight, applications):
    """根据克重 + 应用推断季节"""
    apps = " ".join(applications or [])
    if any(k in apps for k in ["保暖", "外套", "夹棉", "夹克", "工装"]):
        return "秋;冬"
    if "家纺" in apps and (weight or 0) > 250:
        return "秋;冬;四季"
    if (weight or 0) < 200:
        return "春;夏"
    if (weight or 0) < 280:
        return "春;夏;秋"
    return "春;秋;冬"

def infer_features(comp_raw, weight, applications):
    """根据成分 + 克重 + 应用推断特性"""
    out = []
    c = (comp_raw or "").lower()
    if "氨纶" in c or "spandex" in c or "弹" in c:
        out.append("弹力")
    if "再生" in c or "re" in c.lower():
        out.append("再生环保")
    if (weight or 0) >= 300:
        out.append("保暖")
    if (weight or 0) < 200:
        out.append("透气")
    if "亲肤" in (comp_raw or "") or "莫代尔" in (comp_raw or ""):
        out.append("亲肤")
    if not out:
        out.append("挺括")
    return ";".join(out)

def split_apps(apps):
    return ";".join(apps or [])

# ---------- 3. 把现有数据转成表格行 ----------
def build_row(f):
    comp = f.get("composition") or {}
    other_parts = []
    for k, v in comp.items():
        if k not in ("polyester", "cotton", "spandex") and not k.startswith("recycled"):
            other_parts.append(f"{k} {v}%")
    return {
        "面料编号": f.get("code") or f.get("id"),
        "面料名称": f.get("name"),
        "面料图片": "",
        "品类": {"knit": "针织", "woven": "机织", "pu_suede": "PU 绒", "home_textile": "家纺"}.get(f.get("category"), ""),
        "供应商": f.get("supplier") or "",
        "成分描述": f.get("composition_raw") or "",
        "涤纶 %": comp.get("polyester", ""),
        "棉 %": comp.get("cotton", ""),
        "氨纶 %": comp.get("spandex", ""),
        "再生纤维 %": sum(v for k, v in comp.items() if k.startswith("recycled")) if any(k.startswith("recycled") for k in comp) else "",
        "其他成分": ";".join(other_parts),
        "幅宽 cm": f.get("width_cm") or "",
        "克重 g/㎡": f.get("weight_gsm") or (f.get("weight_range") or ""),
        "结构/织法": f.get("weave") or f.get("structure") or "",
        "后整理": f.get("finish") or "",
        "阻燃等级": f.get("fr_standard") or "",
        "RMB 价格 元/m": f.get("price_rmb_per_m") or "",
        "FOB 价格 USD/m": f.get("fob_usd_per_m") or "",
        "起订量 MOQ": f.get("moq") or "",
        "适用季节": infer_season(f.get("weight_gsm"), f.get("applications")),
        "推荐款式": split_apps(f.get("applications") or f.get("tags")),
        "特性标签": infer_features(f.get("composition_raw"), f.get("weight_gsm"), f.get("applications")),
        "卖点文案": "",  # AI 生成
        "相似款链接": "",  # AI 关联
        "备注": f.get("texture") or f.get("edge") or "",
        "__来源文件": f.get("source_file", ""),
        "__来源行号": f.get("source_row", ""),
        "__导入时间": "",  # 导入时自动填
    }

# ---------- 4. 生成模板 ----------
def build_template(path):
    wb = Workbook()

    # Sheet 1: 模板（导入用）
    ws = wb.active
    ws.title = "模板_导入用"
    headers = [c[0] for c in COLUMNS] + [c[0] for c in META_COLUMNS]

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2E5BFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # 1 条样例（用华瑞第一条做示例）
    sample = {
        "面料编号": "ML-EXAMPLE-001",
        "面料名称": "双面弹力摇粒绒（示例，请删除此行后批量导入）",
        "面料图片": "",
        "品类": "针织",
        "供应商": "常熟市华瑞针纺织",
        "成分描述": "94%涤纶 + 6%氨纶",
        "涤纶 %": 94, "棉 %": 0, "氨纶 %": 6, "再生纤维 %": 0, "其他成分": "",
        "幅宽 cm": 165, "克重 g/㎡": 380,
        "结构/织法": "摇粒绒复合",
        "后整理": "磨毛",
        "阻燃等级": "",
        "RMB 价格 元/m": 35, "FOB 价格 USD/m": "", "起订量 MOQ": 500,
        "适用季节": "秋;冬",
        "推荐款式": "外套;卫衣;家居服",
        "特性标签": "保暖;弹力",
        "卖点文案": "【AI 自动生成示例】380g 高克重双面摇粒绒，含 6% 氨纶提供微弹，适合秋冬保暖外套、卫衣和家居服。摇粒绒面蓬松亲肤，可替代传统抓绒，提升成衣档次。",
        "相似款链接": "",
        "备注": "示例行，导入前请删除",
        "__来源文件": "示例", "__来源行号": "", "__导入时间": "",
    }
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=sample.get(h, "")).border = border

    # 列宽
    widths = {"面料编号": 14, "面料名称": 22, "面料图片": 18, "品类": 8, "供应商": 18,
              "成分描述": 18, "涤纶 %": 7, "棉 %": 7, "氨纶 %": 7, "再生纤维 %": 9, "其他成分": 16,
              "幅宽 cm": 8, "克重 g/㎡": 9, "结构/织法": 12, "后整理": 14, "阻燃等级": 14,
              "RMB 价格 元/m": 12, "FOB 价格 USD/m": 12, "起订量 MOQ": 10,
              "适用季节": 12, "推荐款式": 18, "特性标签": 16, "卖点文案": 50, "相似款链接": 22,
              "备注": 20, "__来源文件": 28, "__来源行号": 10, "__导入时间": 12}
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 12)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 60

    # 冻结首行
    ws.freeze_panes = "A2"

    # Sheet 2: 字典
    ws2 = wb.create_sheet("字典")
    ws2.cell(row=1, column=1, value="字段").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="可选值（用分号或换行分隔）").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    r = 2
    for k, vs in DICTS.items():
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=r, column=2, value="; ".join(vs))
        r += 1
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 80

    # Sheet 3: 字段说明
    ws3 = wb.create_sheet("字段说明")
    ws3.cell(row=1, column=1, value="字段名").font = header_font
    ws3.cell(row=1, column=2, value="钉钉类型").font = header_font
    ws3.cell(row=1, column=3, value="必填").font = header_font
    ws3.cell(row=1, column=4, value="说明").font = header_font
    ws3.cell(row=1, column=5, value="示例").font = header_font
    for c in range(1, 6):
        ws3.cell(row=1, column=c).fill = header_fill
    for i, (n, t, req, desc, ex) in enumerate(COLUMNS, 2):
        ws3.cell(row=i, column=1, value=n)
        ws3.cell(row=i, column=2, value=t)
        ws3.cell(row=i, column=3, value=req)
        ws3.cell(row=i, column=4, value=desc)
        ws3.cell(row=i, column=5, value=ex)
    offset = len(COLUMNS) + 2
    for i, (n, t, desc) in enumerate(META_COLUMNS, offset):
        ws3.cell(row=i, column=1, value=n)
        ws3.cell(row=i, column=2, value=t)
        ws3.cell(row=i, column=3, value="")
        ws3.cell(row=i, column=4, value=desc)
        ws3.cell(row=i, column=5, value="")
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 6
    ws3.column_dimensions["D"].width = 40
    ws3.column_dimensions["E"].width = 22

    wb.save(path)
    print(f"✅ 模板已生成：{path}")

# ---------- 5. 生成预转数据 ----------
def build_data_xlsx(json_path, xlsx_path, csv_path):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    rows = [build_row(f) for f in data["fabrics"]]

    # CSV
    headers = [c[0] for c in COLUMNS] + [c[0] for c in META_COLUMNS]
    import csv
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f"✅ CSV 已生成：{csv_path}（{len(rows)} 条）")

    # XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "面料数据"
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E5BFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ri, row in enumerate(rows, 2):
        for col, h in enumerate(headers, 1):
            ws.cell(row=ri, column=col, value=row.get(h, ""))
    # 列宽
    widths = {"面料编号": 16, "面料名称": 24, "面料图片": 14, "品类": 8, "供应商": 18,
              "成分描述": 18, "涤纶 %": 7, "棉 %": 7, "氨纶 %": 7, "再生纤维 %": 9, "其他成分": 16,
              "幅宽 cm": 8, "克重 g/㎡": 9, "结构/织法": 12, "后整理": 14, "阻燃等级": 14,
              "RMB 价格 元/m": 12, "FOB 价格 USD/m": 12, "起订量 MOQ": 10,
              "适用季节": 12, "推荐款式": 18, "特性标签": 16, "卖点文案": 30, "相似款链接": 18,
              "备注": 18, "__来源文件": 30, "__来源行号": 10, "__导入时间": 12}
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 12)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # 数据校验（下拉）- 简版，只对品类和供应商做下拉
    dv_cat = DataValidation(type="list", formula1='"针织,机织,PU 绒,家纺"', allow_blank=True)
    dv_sup = DataValidation(type="list", formula1='"常熟市华瑞针纺织,中涛 / 三时,万泰,home_fr,3S AVVA"', allow_blank=True)
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_sup)
    last_row = len(rows) + 1
    dv_cat.add(f"D2:D{last_row}")
    dv_sup.add(f"E2:E{last_row}")

    wb.save(xlsx_path)
    print(f"✅ XLSX 已生成：{xlsx_path}（{len(rows)} 条）")

if __name__ == "__main__":
    os.makedirs(OUT_DIR, exist_ok=True)
    build_template(os.path.join(OUT_DIR, "面料推荐_钉钉AI表格_导入模板.xlsx"))
    build_data_xlsx(
        SRC_JSON,
        os.path.join(OUT_DIR, "面料推荐_预转数据_86条.xlsx"),
        os.path.join(OUT_DIR, "面料推荐_预转数据_86条.csv"),
    )
    print("\n全部完成 ✅")
